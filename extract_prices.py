# -*- coding: utf-8 -*-
"""
綠源科技 — 成本資料庫萃取器
讀取 成本分析 + 成本分析表 兩個資料夾中所有 Excel 檔案
輸出 price_db.json 供 price-query.html 使用
"""
import os, sys, json, re
from datetime import datetime, date
import openpyxl

BASE = r"C:\Users\user\Desktop"
OUT  = r"C:\Users\user\Desktop\claude\price_db.json"

records = []   # 所有擷取的價格記錄

# ── 工具函式 ──────────────────────────────────────────────────────────
def to_str(v):
    if v is None: return ""
    if isinstance(v, (datetime, date)): return v.strftime("%Y-%m-%d")
    return str(v).strip()

def to_num(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v) if v else None
    try: return float(str(v).replace(",",""))
    except: return None

def parse_date_from_filename(fname):
    """從檔名抓 YYYYMMDD 或 YYYMMDD(民國) 日期"""
    m = re.search(r'(\d{8})', fname)
    if m:
        s = m.group(1)
        if s[:3] in ("111","112","113","114","115"):   # 民國
            yr = int(s[:3]) + 1911
            return f"{yr}-{s[3:5]}-{s[5:7]}"
        if s[:4] in [str(y) for y in range(2020,2030)]:
            return f"{s[:4]}-{s[4:6]}-{s[6:8]}"
    m = re.search(r'(\d{6})', fname)
    if m:
        s = m.group(1)
        if s[:3] in ("111","112","113","114","115"):
            yr = int(s[:3]) + 1911
            return f"{yr}-{s[3:5]}-01"
    return ""

def add(category, item, spec, unit, qty, cost_price, sell_price,
        project, date_str, supplier, source_file, note=""):
    if not item: return
    cost_price = to_num(cost_price)
    sell_price = to_num(sell_price)
    qty        = to_num(qty)
    if cost_price is None and sell_price is None: return
    records.append({
        "category":   category,
        "item":       item,
        "spec":       spec or "",
        "unit":       unit or "",
        "qty":        qty,
        "cost_price": cost_price,
        "sell_price": sell_price,
        "project":    project or "",
        "date":       date_str or "",
        "supplier":   supplier or "",
        "source":     os.path.basename(source_file),
        "note":       note or "",
    })

# ═══════════════════════════════════════════════════════════════════════
# 格式 A：成本分析總表（欄位 項目/內容/單位/數量/單價/複價/說明/COST單價/COST複價/毛利率）
# ═══════════════════════════════════════════════════════════════════════
def extract_cost_analysis_sheet(ws, sheet_name, filepath):
    case_name = sheet_name.strip()
    date_str  = ""
    header_row = None

    # 找 Date 行 和 表頭
    for row in ws.iter_rows(values_only=True):
        row = list(row)
        row_str = " ".join(to_str(v) for v in row)
        if "Date" in row_str or "日期" in row_str:
            # 找日期值
            for v in row:
                if isinstance(v, (datetime, date)):
                    date_str = v.strftime("%Y-%m-%d")
                elif isinstance(v, str):
                    m = re.search(r'(\d{4}[-/]\d{1,2}[-/]\d{1,2})', v)
                    if m: date_str = m.group(1).replace("/","-")
            if not date_str:
                m = re.search(r'CASE[：:](.+)', row_str)
                if m: case_name = m.group(1).strip() or case_name
        # 找表頭行（含「項目」「內容」「單位」「數量」）
        if any(to_str(v) in ("項目","內容","品名") for v in row) and \
           any(to_str(v) in ("單位","數量") for v in row):
            header_row = row
            break

    if header_row is None: return
    # 解析欄位索引
    hdr = [to_str(v).lower() for v in header_row]
    def ci(*keys):
        for k in keys:
            for i,h in enumerate(hdr):
                if k in h: return i
        return None
    c_item   = ci("內容","品名","項目")
    c_unit   = ci("單位")
    c_qty    = ci("數量")
    c_sell   = ci("單價")
    c_cost   = ci("cost單價","cost","進貨","採購")
    c_note   = ci("說明","備註","note")

    if c_item is None: return

    found_header = False
    for row in ws.iter_rows(values_only=True):
        row = list(row)
        row_str = " ".join(to_str(v) for v in row)
        if not found_header:
            if any(to_str(v) in ("內容","品名") for v in row):
                found_header = True
            continue
        item_val = to_str(row[c_item]) if c_item < len(row) else ""
        if not item_val or item_val in ("項目","一","二","三","四","五","六","七","八","九","十"):
            continue
        if re.match(r'^[一二三四五六七八九十]$', item_val): continue

        add(
            category   = "成本分析",
            item       = item_val,
            spec       = "",
            unit       = to_str(row[c_unit])  if c_unit  and c_unit  < len(row) else "",
            qty        = row[c_qty]   if c_qty   and c_qty   < len(row) else None,
            cost_price = row[c_cost]  if c_cost  and c_cost  < len(row) else None,
            sell_price = row[c_sell]  if c_sell  and c_sell  < len(row) else None,
            project    = case_name,
            date_str   = date_str,
            supplier   = "",
            source_file= filepath,
            note       = to_str(row[c_note]) if c_note and c_note < len(row) else ""
        )

# ═══════════════════════════════════════════════════════════════════════
# 格式 B：訂單整理版（分類/合計/單據號碼/廠商簡稱/產品編號/品名規格/數量/單價/金額）
# ═══════════════════════════════════════════════════════════════════════
def extract_order_sheet(ws, filepath):
    fname    = os.path.basename(filepath)
    project  = re.search(r'訂單\d+-(.+?)_', fname)
    project  = project.group(1) if project else fname.replace(".xlsx","")
    date_str = parse_date_from_filename(fname)
    hdr_row  = None

    for row in ws.iter_rows(values_only=True):
        row = list(row)
        if any(to_str(v) in ("品名規格","品名","規格") for v in row):
            hdr_row = row
            break
    if hdr_row is None: return

    hdr = [to_str(v) for v in hdr_row]
    def ci(*keys):
        for k in keys:
            for i,h in enumerate(hdr):
                if k in h: return i
        return None
    c_cat  = ci("分類")
    c_code = ci("單據號碼","單據")
    c_sup  = ci("廠商簡稱","廠商")
    c_item = ci("品名規格","品名","規格")
    c_qty  = ci("數量")
    c_up   = ci("單價")
    c_amt  = ci("金額")

    if c_item is None: return

    found = False
    current_cat = ""
    for row in ws.iter_rows(values_only=True):
        row = list(row)
        if not found:
            if any(to_str(v) in ("品名規格","品名","規格") for v in row):
                found = True
            continue
        item_val = to_str(row[c_item]) if c_item < len(row) else ""
        if not item_val: continue

        cat_val = to_str(row[c_cat]) if c_cat is not None and c_cat < len(row) else ""
        if cat_val: current_cat = cat_val

        # 從品名規格拆出日期（有時格式是 品名/案場/廠商/有效期）
        parts = [p.strip() for p in item_val.split("/") if p.strip()]
        item_clean = parts[0] if parts else item_val

        # 嘗試從單據號碼中取日期
        code_val = to_str(row[c_code]) if c_code is not None and c_code < len(row) else ""
        item_date = parse_date_from_filename(code_val) or date_str

        add(
            category   = current_cat or "採購",
            item       = item_clean,
            spec       = "/".join(parts[1:]) if len(parts)>1 else "",
            unit       = "",
            qty        = row[c_qty] if c_qty and c_qty < len(row) else None,
            cost_price = row[c_up]  if c_up  and c_up  < len(row) else None,
            sell_price = None,
            project    = project,
            date_str   = item_date,
            supplier   = to_str(row[c_sup]) if c_sup and c_sup < len(row) else "",
            source_file= filepath,
            note       = ""
        )

# ═══════════════════════════════════════════════════════════════════════
# 格式 C：跨案場比較表（行=項目，列=案場，格式 項目/數量/單價/總價 循環）
# ═══════════════════════════════════════════════════════════════════════
def extract_cross_project_sheet(ws, filepath):
    fname    = os.path.basename(filepath)
    date_str = parse_date_from_filename(fname)
    rows_data = list(ws.iter_rows(values_only=True))
    if len(rows_data) < 2: return

    # Row 0: project names (每3欄一組 數量/單價/總價)
    # Row 1: 項目, 數量, 單價, 總價, 數量, 單價, 總價...
    header0 = [to_str(v) for v in rows_data[0]]
    header1 = [to_str(v) for v in rows_data[1]]

    # 找項目欄
    item_col = None
    for i, h in enumerate(header1):
        if "項目" in h or "品名" in h or h == "":
            if item_col is None:
                item_col = i

    # 找案場和其欄位索引
    projects = {}
    proj_name = ""
    for i, h in enumerate(header0):
        if h and h not in ("", "None") and "項目" not in h:
            proj_name = h
        if "單價" in header1[i] if i < len(header1) else False:
            if proj_name:
                qty_col  = i - 1 if i > 0 else None
                projects[proj_name] = {"qty": qty_col, "price": i}

    if not projects: return

    for row in rows_data[2:]:
        row = list(row)
        item_val = to_str(row[item_col]) if item_col is not None and item_col < len(row) else ""
        if not item_val or item_val == "None": continue
        try:
            # Skip if item starts with digit category number
            if re.match(r'^\d', item_val) and len(item_val) < 5: continue
        except: pass

        for proj, cols in projects.items():
            pc = cols.get("price")
            qc = cols.get("qty")
            if pc is None or pc >= len(row): continue
            price_val = to_num(row[pc])
            qty_val   = to_num(row[qc]) if qc is not None and qc < len(row) else None
            if price_val:
                add(
                    category   = "土建工程",
                    item       = item_val,
                    spec       = "",
                    unit       = "",
                    qty        = qty_val,
                    cost_price = price_val,
                    sell_price = None,
                    project    = proj,
                    date_str   = date_str,
                    supplier   = "",
                    source_file= filepath,
                )

# ═══════════════════════════════════════════════════════════════════════
# 格式 D：通用 Excel — 找含「單價」欄位的工作表
# ═══════════════════════════════════════════════════════════════════════
def extract_generic(ws, filepath, category=""):
    fname    = os.path.basename(filepath)
    date_str = parse_date_from_filename(fname)
    project  = fname.replace(".xlsx","").replace(".xls","")

    hdr_row = None
    for row in ws.iter_rows(values_only=True):
        row = list(row)
        row_str = " ".join(to_str(v) for v in row)
        if "單價" in row_str and ("品名" in row_str or "項目" in row_str or "內容" in row_str):
            hdr_row = row
            break
    if hdr_row is None: return

    hdr = [to_str(v) for v in hdr_row]
    def ci(*keys):
        for k in keys:
            for i,h in enumerate(hdr):
                if k in h: return i
        return None
    c_item = ci("品名","項目","內容","規格","說明")
    c_qty  = ci("數量","Qty","QTY")
    c_up   = ci("單價","Unit Price","unit price")
    c_unit = ci("單位","Unit","unit")
    c_note = ci("備註","說明","Note")
    c_sup  = ci("廠商","供應商","Vendor")

    if c_item is None or c_up is None: return

    found = False
    for row in ws.iter_rows(values_only=True):
        row = list(row)
        if not found:
            if any(to_str(v) in ("品名","項目","內容","規格") for v in row):
                found = True
            continue
        item_val = to_str(row[c_item]) if c_item < len(row) else ""
        if not item_val: continue

        add(
            category   = category or "其他",
            item       = item_val,
            spec       = "",
            unit       = to_str(row[c_unit]) if c_unit and c_unit < len(row) else "",
            qty        = row[c_qty] if c_qty and c_qty < len(row) else None,
            cost_price = row[c_up]  if c_up  and c_up  < len(row) else None,
            sell_price = None,
            project    = project,
            date_str   = date_str,
            supplier   = to_str(row[c_sup]) if c_sup and c_sup < len(row) else "",
            source_file= filepath,
            note       = to_str(row[c_note]) if c_note and c_note < len(row) else ""
        )

# ═══════════════════════════════════════════════════════════════════════
# 主流程：掃描所有 Excel 檔案
# ═══════════════════════════════════════════════════════════════════════
FOLDER_CATEGORY = {
    "電線電纜":      "電線電纜",
    "鋼構支架類":    "鋼構支架",
    "高低壓盤及變壓器": "高低壓盤/變壓器",
    "土建":          "土建工程",
    "線槽":          "線槽",
    "監控系統":      "監控系統",
    "勞安":          "勞安/保險",
    "雜項工程":      "雜項工程",
    "跑件":          "跑件/設計",
    "表前儲能":      "儲能工程",
    "大型項目":      "大型項目",
    "表後":          "表後系統",
    "成本分析":      "成本分析",
}

def get_category(filepath):
    parts = filepath.replace("\\","/").split("/")
    for p in parts:
        for k, v in FOLDER_CATEGORY.items():
            if k in p: return v
    return "其他"

def process_file(filepath):
    cat = get_category(filepath)
    fname = os.path.basename(filepath).lower()
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"  ⚠ 無法讀取 {os.path.basename(filepath)}: {e}", file=sys.stderr)
        return

    # 選擇適合的萃取格式
    if "成本分析總表" in os.path.basename(filepath):
        for sname in wb.sheetnames:
            ws = wb[sname]
            extract_cost_analysis_sheet(ws, sname, filepath)

    elif "訂單" in os.path.basename(filepath):
        for ws in [wb[s] for s in wb.sheetnames]:
            extract_order_sheet(ws, filepath)

    elif "儲能成本分析" in os.path.basename(filepath):
        for ws in [wb[s] for s in wb.sheetnames]:
            extract_cross_project_sheet(ws, filepath)

    elif any(k in os.path.basename(filepath) for k in ("信邦","採購清單","炳達","成本分析")):
        ws = wb.active
        extract_cost_analysis_sheet(ws, wb.sheetnames[0], filepath)

    else:
        # 通用格式嘗試
        for sname in wb.sheetnames:
            ws = wb[sname]
            before = len(records)
            extract_generic(ws, filepath, cat)
            if len(records) > before:
                break
        else:
            # Try cost analysis format on first sheet
            ws = wb.active
            extract_cost_analysis_sheet(ws, wb.sheetnames[0], filepath)

total = 0
for folder in ["成本分析", "成本分析表"]:
    folder_path = os.path.join(BASE, folder)
    for root, dirs, files in os.walk(folder_path):
        for fname in files:
            if fname.startswith("~$"): continue
            if not (fname.endswith(".xlsx") or fname.endswith(".xls")): continue
            fpath = os.path.join(root, fname)
            before = len(records)
            process_file(fpath)
            n = len(records) - before
            print(f"  {n:3d} 筆  {os.path.relpath(fpath, BASE)}", file=sys.stderr)
            total += n

print(f"\n✅ 共萃取 {total} 筆價格記錄", file=sys.stderr)

# 去除空值清洗
clean = [r for r in records if r["item"] and (r["cost_price"] or r["sell_price"])]

# 輸出 JSON
with open(OUT, "w", encoding="utf-8") as f:
    json.dump({
        "generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "total": len(clean),
        "records": clean
    }, f, ensure_ascii=False, indent=2)

print(f"✅ 輸出 {len(clean)} 筆 → {OUT}", file=sys.stderr)
